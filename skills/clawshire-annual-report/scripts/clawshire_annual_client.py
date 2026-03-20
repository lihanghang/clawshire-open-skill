#!/usr/bin/env python3
"""
ClawShire 年报数据 API 客户端
北交所上市公司年度报告结构化提取结果查询工具

用法：
    python clawshire_annual_client.py stock 833359 --year 2024
    python clawshire_annual_client.py list --year 2024
    python clawshire_annual_client.py met-link "https://www.bse.cn/..."
    python clawshire_annual_client.py api-key-info

    # 导出表格
    python clawshire_annual_client.py list --year 2024 --output csv
    python clawshire_annual_client.py list --year 2024 --output excel
    python clawshire_annual_client.py stock 833359 --output excel

获取 API Key：登录控制台 https://clawshire.cn 手动创建
设置环境变量：export CLAWSHIRE_API_KEY="sk-xxxxxxxx"
"""

import argparse
import csv
import io
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from typing import Any

BASE_URL = "https://api.clawshire.cn"
CONSOLE_URL = "https://clawshire.cn"
INFOTYPE = "年报"
DEFAULT_PAGE_SIZE = 20
MAX_RETRIES = 3
TIMEOUT = 30

# 北交所证券代码前缀
BSE_PREFIXES = ("4", "8", "9")


def _request(
    method: str,
    path: str,
    params: dict | None = None,
    api_key: str | None = None,
) -> dict:
    """执行 HTTP 请求，内置重试逻辑。"""
    url = f"{BASE_URL}{path}"
    if params:
        url = f"{url}?{urllib.parse.urlencode(params)}"

    headers = {"Content-Type": "application/json", "Accept": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    for attempt in range(MAX_RETRIES):
        try:
            req = urllib.request.Request(url, headers=headers, method=method)
            with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            error_body = {}
            try:
                error_body = json.loads(e.read().decode("utf-8"))
            except Exception:
                pass
            print(
                json.dumps(
                    {"error": f"HTTP {e.code}", "detail": error_body},
                    ensure_ascii=False,
                    indent=2,
                ),
                file=sys.stderr,
            )
            sys.exit(1)
        except urllib.error.URLError as e:
            if attempt < MAX_RETRIES - 1:
                wait = 2 ** attempt
                print(f"请求失败（{e.reason}），{wait}s 后重试...", file=sys.stderr)
                time.sleep(wait)
            else:
                print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
                sys.exit(1)

    sys.exit(1)


def _require_api_key(args_key: str | None) -> str:
    key = args_key or os.environ.get("CLAWSHIRE_API_KEY")
    if not key:
        print(
            f"错误：未设置 CLAWSHIRE_API_KEY 环境变量。\n"
            f"请登录控制台创建 API Key：{CONSOLE_URL}\n"
            f"创建后执行：export CLAWSHIRE_API_KEY='sk-xxxxxxxx'",
            file=sys.stderr,
        )
        sys.exit(1)
    return key


def _year_to_date_range(year: int) -> tuple[str, str]:
    """将年份转换为年报披露的日期范围（次年 1-1 ~ 次年 6-30）。"""
    return f"{year + 1}-01-01", f"{year + 1}-06-30"


def cmd_stock_annual(args: argparse.Namespace) -> dict:
    """查询指定北交所公司的年报提取结果。"""
    sec_code = args.sec_code
    if not sec_code.startswith(BSE_PREFIXES):
        print(
            f"警告：{sec_code} 不是北交所证券代码（应以 4/8/9 开头），将继续查询但可能无结果。",
            file=sys.stderr,
        )

    api_key = _require_api_key(args.api_key)
    params: dict[str, Any] = {
        "infotype": INFOTYPE,
        "page": args.page,
        "page_size": args.page_size,
    }
    if args.year:
        start, end = _year_to_date_range(args.year)
        params["start_date"] = start
        params["end_date"] = end
    elif args.start_date:
        params["start_date"] = args.start_date
        if args.end_date:
            params["end_date"] = args.end_date

    return _request("GET", f"/api/v1/stock/{sec_code}/announcements", params=params, api_key=api_key)


def cmd_list_annual(args: argparse.Namespace) -> dict:
    """查询北交所全市场年报提取结果。"""
    api_key = _require_api_key(args.api_key)
    params: dict[str, Any] = {
        "infotype": INFOTYPE,
        "page": args.page,
        "page_size": args.page_size,
    }
    if args.year:
        start, end = _year_to_date_range(args.year)
        params["start_date"] = start
        params["end_date"] = end
    elif args.start_date:
        params["start_date"] = args.start_date
        if args.end_date:
            params["end_date"] = args.end_date

    return _request("GET", "/api/v1/announcements", params=params, api_key=api_key)


def cmd_met_link(args: argparse.Namespace) -> dict:
    """根据年报原文链接查询提取结果。"""
    api_key = _require_api_key(args.api_key)
    params: dict[str, Any] = {
        "met_link": args.link,
        "infotype": INFOTYPE,
    }
    return _request("GET", "/api/v1/met_link", params=params, api_key=api_key)


def cmd_api_key_info(args: argparse.Namespace) -> dict:
    """查看当前 API Key 信息。"""
    api_key = _require_api_key(args.api_key)
    return _request("GET", "/api/v1/api-key/info", api_key=api_key)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="ClawShire 年报数据 API 客户端（北交所）",
        epilog=f"获取 API Key：登录控制台 {CONSOLE_URL} 手动创建",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--api-key", help="API Key（默认读取 CLAWSHIRE_API_KEY 环境变量）")

    sub = parser.add_subparsers(dest="command", required=True)

    # stock
    p_stock = sub.add_parser("stock", help="按证券代码查询年报（北交所 4/8/9 开头）")
    p_stock.add_argument("sec_code", help="六位证券代码，如 833359")
    p_stock.add_argument("--year", type=int, help="报告年份，如 2024（优先于 start/end-date）")
    p_stock.add_argument("--start-date", default="", help="开始日期 YYYY-MM-DD")
    p_stock.add_argument("--end-date", default="", help="结束日期 YYYY-MM-DD")
    p_stock.add_argument("--page", type=int, default=1)
    p_stock.add_argument("--page-size", type=int, default=DEFAULT_PAGE_SIZE)
    p_stock.add_argument("--output", choices=["csv", "excel"], help="导出格式（csv / excel）")

    # list
    p_list = sub.add_parser("list", help="查询北交所全市场年报")
    p_list.add_argument("--year", type=int, help="报告年份，如 2024")
    p_list.add_argument("--start-date", default="", help="开始日期 YYYY-MM-DD")
    p_list.add_argument("--end-date", default="", help="结束日期 YYYY-MM-DD")
    p_list.add_argument("--page", type=int, default=1)
    p_list.add_argument("--page-size", type=int, default=DEFAULT_PAGE_SIZE)
    p_list.add_argument("--output", choices=["csv", "excel"], help="导出格式（csv / excel）")

    # met-link
    p_link = sub.add_parser("met-link", help="按年报原文链接查询")
    p_link.add_argument("link", help="年报原文 URL")
    p_link.add_argument("--output", choices=["csv", "excel"], help="导出格式（csv / excel）")

    # api-key-info
    sub.add_parser("api-key-info", help="查看当前 API Key 信息")

    return parser


COMMAND_MAP = {
    "stock": cmd_stock_annual,
    "list": cmd_list_annual,
    "met-link": cmd_met_link,
    "api-key-info": cmd_api_key_info,
}


def _flatten_extracted(info: Any) -> dict:
    """将 extracted_info 扁平化为单层 dict。"""
    if not info:
        return {}
    if isinstance(info, list):
        info = info[0] if info else {}
    if not isinstance(info, dict):
        return {}
    flat = {}
    for k, v in info.items():
        if isinstance(v, dict):
            for sk, sv in v.items():
                flat[f"{k}_{sk}"] = str(sv) if sv is not None else ""
        elif isinstance(v, list):
            flat[k] = "; ".join(str(i) for i in v)
        else:
            flat[k] = str(v) if v is not None else ""
    return flat


def _to_rows(result: dict) -> list[dict]:
    """将 API 返回结果转为扁平行列表。"""
    items = result.get("data", result.get("items", []))
    if isinstance(result.get("data"), dict):
        # 单条（met-link 返回）
        items = [result["data"]]

    rows = []
    for item in items:
        base = {
            "证券代码": item.get("sec_code", ""),
            "公司名称": item.get("sec_name", ""),
            "公告标题": item.get("announcement_title", ""),
            "公告日期": str(item.get("announcement_time", ""))[:10],
            "原文链接": item.get("met_link", item.get("adjunct_url", "")),
        }
        extracted = _flatten_extracted(item.get("extracted_info"))
        base.update(extracted)
        rows.append(base)
    return rows


def _export_csv(rows: list[dict], path: str) -> None:
    if not rows:
        print("无数据可导出", file=sys.stderr)
        return
    keys = list(dict.fromkeys(k for r in rows for k in r))
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"已导出 CSV：{path}（{len(rows)} 条）")


def _export_excel(rows: list[dict], path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("需要安装 openpyxl：pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not rows:
        print("无数据可导出", file=sys.stderr)
        return

    keys = list(dict.fromkeys(k for r in rows for k in r))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "年报数据"

    # 表头样式
    header_fill = PatternFill("solid", fgColor="1A6FC4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, key in enumerate(keys, 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, row in enumerate(rows, 2):
        for col, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col, value=row.get(key, ""))

    # 自动列宽（最大 40）
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(path)
    print(f"已导出 Excel：{path}（{len(rows)} 条）")


def _default_filename(args: argparse.Namespace, ext: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    sec = getattr(args, "sec_code", None)
    year = getattr(args, "year", None)
    parts = ["annual"]
    if sec:
        parts.append(sec)
    if year:
        parts.append(str(year))
    parts.append(ts)
    return "_".join(parts) + f".{ext}"


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    handler = COMMAND_MAP.get(args.command)
    if not handler:
        parser.print_help()
        sys.exit(1)

    result = handler(args)

    output = getattr(args, "output", None)
    if output == "csv":
        rows = _to_rows(result)
        _export_csv(rows, _default_filename(args, "csv"))
    elif output == "excel":
        rows = _to_rows(result)
        _export_excel(rows, _default_filename(args, "xlsx"))
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
